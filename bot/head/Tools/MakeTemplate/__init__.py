import pytz
import os
import openpyxl
from pytz import timezone
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from bot.head.Tools.MakeTemplate.appends import listas


class MakeXlsx:

    def __init__(self, type: str, bot: str):
        """
        
        #### type: Tipo da planilha (sucesso, erro)
        #### bot: o sistema que está sendo executado a automação Ex.: PROJUDI, ESAJ, ELAW, ETC.
            
        """
        self.bot = bot
        self.type = type
        self.listas = listas()
        pass
    
    def make_output(self, path_template: str):

        # Criar um novo workbook e uma planilha
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Resultados", 0)
        sheet = workbook.active

        # Cabeçalhos iniciais
        cabecalhos = ["NUMERO_PROCESSO"]
        list_to_append = []
        
        itens_append = self.listas(f"{self.bot}_{self.type}")
        if itens_append:
            list_to_append.extend(itens_append)
        
        elif not itens_append:
            itens_append = self.listas(self.type)
            if itens_append:
                list_to_append.extend(itens_append)

        # Definir estilo
        my_red = openpyxl.styles.colors.Color(rgb='A6A6A6')
        my_fill = PatternFill(patternType='solid', fgColor=my_red)
        bold_font = Font(name='Times New Roman', italic=True)

        # Escrever os cabeçalhos na primeira linha da planilha e aplicar estilo
        for pos, coluna in enumerate(cabecalhos):
            item = sheet.cell(row=1, column=pos+1, value=coluna.upper())
            item.font = bold_font
            item.fill = my_fill

        # Ajustar a largura das colunas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # Salvar o workbook no caminho especificado
        workbook.save(path_template)

        return True

# from datetime import datetime
# namefile = f'Busca Esaj Capa - {datetime.now(pytz.timezone('Etc/GMT+4')).strftime("%d-%m-%y")}.xlsx'
# localtosave = f"{os.path.join(pathlib.Path(__file__).parent.resolve(), namefile)}"
# args = 'esaj_emite_guia'

# MakeXlsx().make_template(args, localtosave)
